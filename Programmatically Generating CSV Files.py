from datetime import datetime
from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference


# johten7331
# Graphing Dynamically Generated Data


# Get the folder where this Python file is located.
BASE_DIR = Path(__file__).resolve().parent

# Store the CSV and Excel files in the same folder as this program.
CSV_FILE = BASE_DIR / "ZooData.csv"
EXCEL_FILE = BASE_DIR / "final.xlsx"


# Converts a Fahrenheit temperature to Celsius.
def convertData(data):
    return (data - 32) * 5 / 9


# Opens or creates a CSV file and appends data to it.
def insertData(path, date, fahrenheit, celsius):
    try:
        with open(path, "a", newline="") as file:
            writer = csv.writer(file)
            writer.writerow([date, fahrenheit, celsius])

        return True

    except Exception as error:
        print("Error writing to file:", error)
        return False


# Reads and displays the contents of a CSV file.
def viewData(path):
    try:
        with open(path, "r", newline="") as file:
            print("\nReading data from:", path)
            print()

            reader = csv.reader(file)

            for row in reader:
                if len(row) >= 3:
                    print(
                        "Date:",
                        row[0],
                        "| Fahrenheit:",
                        row[1],
                        "| Celsius:",
                        row[2]
                    )

    except FileNotFoundError:
        print("Error: The CSV file does not exist yet.")

    except Exception as error:
        print("Error reading file:", error)


# Gets temperature data from the user, converts it, and saves it to the CSV file.
def getInput():
    try:
        entries = int(input("How many entries are you inputting? "))

        for i in range(entries):
            print("\nEntry", i + 1)

            date = input("Enter a date: ")

            temperature = float(
                input("Enter the highest temp for the inputted date: ")
            )

            convertedTemperature = convertData(temperature)

            if insertData(
                CSV_FILE,
                date,
                temperature,
                convertedTemperature
            ):
                print(
                    "The following data was saved at",
                    datetime.now()
                )

                print(
                    date,
                    temperature,
                    convertedTemperature
                )

    except ValueError:
        print("Error: Please enter a valid number.")

    except Exception as error:
        print("Error:", error)


# Creates a chart using temperature data stored in a CSV file.
# Arguments:
# path (string or Path): Path to the CSV data file.
# chartType (string): Type of chart to create, either "line" or "bar".
# Return value: None.
def createChart(path, chartType):
    try:
        print("\nChoose the data source for the report")
        print("1 Fahrenheit")
        print("2 Celsius")

        dataChoice = input("Enter your choice: ")

        if dataChoice == "1":
            csvColumn = 1
            columnHeader = "Fahrenheit"
            yAxisLabel = "Temperature (Fahrenheit)"

        elif dataChoice == "2":
            csvColumn = 2
            columnHeader = "Celsius"
            yAxisLabel = "Temperature (Celsius)"

        else:
            print("Error: Invalid data source.")
            return

        dates = []
        temperatures = []

        # Open the CSV file and extract the selected temperature data.
        with open(path, "r", newline="") as file:
            reader = csv.reader(file)

            for row in reader:
                if len(row) >= 3:
                    date = row[0]

                    # Cast the selected temperature value to a float.
                    temperature = float(row[csvColumn])

                    dates.append(date)
                    temperatures.append(temperature)

        if len(dates) == 0:
            print("Error: No data was found in the CSV file.")
            return

        # Create a new Excel workbook.
        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "Temperature Report"

        # Add headers to the worksheet.
        worksheet["A1"] = "Date"
        worksheet["B1"] = columnHeader

        # Add the selected CSV data to the worksheet.
        for i in range(len(dates)):
            worksheet.cell(
                row=i + 2,
                column=1,
                value=dates[i]
            )

            worksheet.cell(
                row=i + 2,
                column=2,
                value=temperatures[i]
            )

        # Create the chart type passed to the function.
        if chartType == "line":
            chart = LineChart()

        elif chartType == "bar":
            chart = BarChart()

        else:
            print("Error: Invalid chart type.")
            return

        # Select the temperature data from the worksheet.
        data = Reference(
            worksheet,
            min_col=2,
            min_row=1,
            max_row=len(temperatures) + 1
        )

        # Select the dates to use as x-axis labels.
        categories = Reference(
            worksheet,
            min_col=1,
            min_row=2,
            max_row=len(dates) + 1
        )

        # Add the temperature data to the chart.
        chart.add_data(
            data,
            titles_from_data=True
        )

        # Use the dates as category labels.
        chart.set_categories(categories)

        # Label the chart axes.
        chart.x_axis.title = "Date"
        chart.y_axis.title = yAxisLabel

        # Make sure the chart axes and labels are displayed.
        chart.x_axis.delete = False
        chart.y_axis.delete = False

        chart.x_axis.tickLblPos = "low"
        chart.y_axis.tickLblPos = "low"

        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickMarkSkip = 1

        chart.x_axis.axPos = "b"
        chart.y_axis.axPos = "l"

        # Give the y-axis a useful range.
        chart.y_axis.scaling.min = min(temperatures) - 10
        chart.y_axis.scaling.max = max(temperatures) + 10

        # Create the required chart title.
        currentDate = datetime.now().strftime("%m/%d/%Y")

        chart.title = "johten7331 " + currentDate

        # There is only one temperature series.
        chart.legend = None

        # Set chart dimensions.
        chart.height = 10
        chart.width = 18

        # Make worksheet columns readable.
        worksheet.column_dimensions["A"].width = 18
        worksheet.column_dimensions["B"].width = 24

        # Add the chart to the worksheet.
        worksheet.add_chart(chart, "D2")

        # Save the Excel workbook.
        workbook.save(EXCEL_FILE)

        print()
        print("Report successfully created as final.xlsx")

    except FileNotFoundError:
        print("Error: The CSV file could not be found.")

    except ValueError:
        print("Error: The CSV file contains invalid temperature data.")

    except PermissionError:
        print("Error: final.xlsx is currently open.")
        print("Close Excel and run the program again.")

    except Exception as error:
        print("Error creating report:", error)


# Prompts the user to select a graph type and calls createChart.
# Argument:
# path (string or Path): Path to the CSV data file.
# Return value: None.
def generateReport(path):
    print("\nChoose a graph type")
    print("1 Line Chart")
    print("2 Bar Chart")

    graphChoice = input("Enter your choice: ")

    if graphChoice == "1":
        createChart(path, "line")

    elif graphChoice == "2":
        createChart(path, "bar")

    else:
        print("Error: Invalid graph type.")


# Displays the main menu and handles the user's selection.
def main():
    print()
    print("johten7331's Spreadsheet Automation Menu")
    print("Choose a number from the following options")
    print()

    menu_options = [
        "Input Data",
        "View Current Data",
        "Generate Report"
    ]

    for i in range(len(menu_options)):
        print(i + 1, menu_options[i])

    choice = input("\nEnter your choice: ")

    if choice == "1":
        print(
            "You selected",
            choice,
            "at",
            datetime.now()
        )

        getInput()

    elif choice == "2":
        print(
            "You selected",
            choice,
            "at",
            datetime.now()
        )

        viewData(CSV_FILE)

    elif choice == "3":
        print(
            "You selected",
            choice,
            "at",
            datetime.now()
        )

        generateReport(CSV_FILE)

    else:
        print("Error: Invalid selection.")


main()